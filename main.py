import os
import os.path
from os import path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import logging
import pandas as pd
from copy import copy
import time
import json
from datetime import timedelta

def date_range (dset=7):
    ##TODO: Need to fix date
    ## if you need a different week use the dset variable
    ddate = datetime.today().toordinal()-(dset+datetime.today().weekday())
    sdate = datetime.fromordinal(ddate).strftime('%m-%d-%Y')
    edate = datetime.fromordinal(ddate+7).strftime('%m-%d-%Y')
    return {'start date':sdate,'end date':edate}

irDir = os.getcwd()+'/input/'
irWorksheet = os.getcwd()+'/irSummary.xlsx'
outputDir = os.getcwd()+'/output/'
dateRange = date_range()
outputName = 'LnW IR Summary '+dateRange['end date']+'.xlsx'

## Iterates through a directory to include each incident report into a list
def iterateIRs():
    def convertDate(value):
        value = datetime.strptime(value, "%m-%d-%y")
        value = value.strftime("%m-%d-%y")
        return value
    irOutputDict= {}
    for root, dirs, reports in os.walk(irDir):
        for report in reports:
            try:
                wb = load_workbook(root+report)
                ws=wb.active
            except:
                ## TODO: Log Error for being unable to read IR
                msg = 'Unable to read IR file. {}'.format(report)
                print(msg)
                pass
            ## TODO: N/A does not show
            try:
                irWorksheetValues = {
            'Facility Name':ws['E3'].value,
            'Technician':ws['E4'].value,
            'Manager':ws['E5'].value,
            'Date of Incident':ws['E6'].value.strftime("%m-%d-%y"),
            'Start Time':ws['E7'].value,
            'End Time and Date':ws['E8'].value,
            'Type of Problem':ws['D10'].value,
            'Description':ws['D12'].value,
            'Devices Involved':ws['D14'].value,
            'Brief Description':ws['A17'].value,
            'Additional Details':ws['A23'].value,
            'Additional Details Mgmt':ws['A28'].value,
            'Type of Incident':ws['D33'].value,
            'Cause':ws['D34'].value,
            'Effect':ws['D35'].value,
            'Lead':ws['D36'].value,
            'Supervisor':ws['D37'].value,
            'Machine Freeze':ws['D40'].value,
            'Communication Error':ws['D41'].value,
            'Tilts and/or Errors':ws['D42'].value,
            'Connectivity':ws['D45'].value,
            'Cash-Out':ws['D46'].value,
            'Reboot Process':ws['D47'].value,
            'Ram Clear Process':ws['D48'].value,
            'EVERI Session History':ws['D51'].value,
            'EVERI Event Log':ws['D52'].value,
            'Machine Event History':ws['D53'].value,
            'Digital Pictues':ws['D54'].value,
            'Surveillance Tape':ws['D55'].value,
            'Revenue Loss':ws['D58'].value,
            'Parts Loss':ws['D59'].value}
                count=0
                for key, value in irWorksheetValues.items():
                    if value == 'nan':
                        ##TODO: Logging needed
                        msg = 'IR {} in column {} has a blank response.'.format(report,key)
                        print(msg)
                    else:
                        pass
                    ##TODO: Track Abbreviations
                    abr = {'Suffolk':'SFK'}
                    try:
                        ir_name = str(irWorksheetValues['Facility Name']+'-'+irWorksheetValues['Date of Incident']+'-'+str(count))
                        msg = ''
                    except AttributeError as e:
                        ##TODO:Logging 'error'
                        msg='Invalid value for date on {}- date is set as {}'.format(report,irWorksheetValues['Date of Incident'])
                        ir_name = str(irWorksheetValues['Facility Name']+'-'+irWorksheetValues['Date of Incident'])
                    count=count+1
                if msg:
                    print(msg)
                irOutputDict.update({ir_name:irWorksheetValues})
                wb.close()
            except:
                print('error')
                pass
    return irOutputDict


def IR_format():
    try:
        ws = wb['Incident Report Summary']
    except KeyError as e:
        print('KeyError - Worksheet Incident Report Summary does not exist')
        print('Skipping process')
        return
    for cell in ws[1]:
        cell.value=''
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    ws['A1'].value = 'Incident Report Summary'
    ws['A1'].font = header_cell_font
    for rows in ws.iter_rows():
        l = []
        [l.append(cell.value) for cell in rows]
        if l == ['Track', 'Location', 'Serial #', 'Device Name', 'Title', 'Cabinet', 'Date', 'Time In', 'Time Playable', 'Total Time Down', 'Tech', 'Comments']:
            for cell in rows:
                cell.font=header_cell_font
                cell.fill = header_bgcolor

def width_format():
    #column width format assigned from previous dims item
    for name in names:
        dims = {}
        for col in wb[name].iter_cols():
            for cell in col:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
        for col, value in dims.items():
            wb[name].column_dimensions[col].width = value+1

def datetime_format():
    #date and time format
    rows = ['Time In', 'Time Out', 'Total Time Down']
    ws= wb['TechLog']
    for row in ws.iter_cols(1, ws.max_column):
        try:
            if row[0].value == 'Date':
                ###number_format is not working properly
                for cell in row:
                    cell.number_format='mm-dd-yy'
            elif row[0].value == rows[0]:
                for cell in row:
                    cell.number_format='h:mm'
            elif row[0].value == rows[1]:
                for cell in row:
                    cell.number_format='h:mm'
            elif row[0].value == rows[2]: 
                for cell in row:
                    cell.number_format='h:mm'
            else:
                pass
        except Exception as e:
            print('an error has occurred: {} with value {} at {}.'.format(e, cell.value, cell.coordinate))
            pass

def write_dict():
    with pd.ExcelWriter(outputName,'xlsxwriter', mode='w') as writer:
        for key, value in iterateIRs():
            value.to_excel(writer, sheet_name=key, index=False)
        writer.save()
        writer.close()

write_dict()